"""
report_generator.py
-------------------
Professional report generation utilities for the Student Performance Dashboard.

Supported formats:
    * PDF (ReportLab) - multi-section professional report with embedded charts
    * Excel (OpenPyXL) - formatted workbook with raw data, statistics, KPIs,
      and correlation matrix
    * CSV - filtered dataset export

All generators accept the *currently filtered* DataFrame so reports reflect
exactly what the user sees on the dashboard.
"""

from __future__ import annotations

import io
import logging
import os
import tempfile
from datetime import datetime
from typing import Any

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.io import to_image

from utils.chart_theme import apply_theme, empty_figure
from utils.data_loader import compute_kpis, filter_dataframe, load_clean_data

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Optional dependency guards
# ---------------------------------------------------------------------------
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Image as RLImage,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:  # noqa: BLE001
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    OPENPYXL_AVAILABLE = True
except ImportError:  # noqa: BLE001
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_pdf_report(
    df: pd.DataFrame,
    filters: dict[str, Any],
    full_df: pd.DataFrame,
    insights: list[str] | None = None,
) -> bytes:
    """Return a complete PDF report as bytes.

    Parameters
    ----------
    df : pd.DataFrame
        The currently filtered dataset.
    filters : dict
        Active filter values (for the dataset summary section).
    full_df : pd.DataFrame
        The complete cleaned dataset (for missing-value context).
    insights : list[str], optional
        Pre-computed insight strings.
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError(
            "ReportLab is required for PDF reports. "
            "Install it with: pip install reportlab"
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Custom styles
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=26,
        leading=32,
        spaceAfter=18,
        textColor=colors.HexColor("#1E293B"),
    )
    heading1_style = ParagraphStyle(
        "ReportH1",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        spaceAfter=10,
        spaceBefore=16,
        textColor=colors.HexColor("#1E293B"),
        borderWidth=0,
        borderColor=colors.HexColor("#E2E8F0"),
        borderPadding=8,
        backColor=colors.HexColor("#F8FAFC"),
    )
    heading2_style = ParagraphStyle(
        "ReportH2",
        parent=styles["Heading2"],
        fontSize=13,
        leading=16,
        spaceAfter=6,
        spaceBefore=12,
        textColor=colors.HexColor("#334155"),
    )
    body_style = ParagraphStyle(
        "ReportBody",
        parent=styles["BodyText"],
        fontSize=10,
        leading=14,
        spaceAfter=6,
        textColor=colors.HexColor("#475569"),
    )
    caption_style = ParagraphStyle(
        "Caption",
        parent=styles["Italic"],
        fontSize=9,
        leading=12,
        alignment=1,
        textColor=colors.HexColor("#64748B"),
    )

    story = []
    pages_info = {"total": 0, "current": 0}

    def add_page_number(canvas, doc_obj):
        pages_info["current"] += 1
        pages_info["total"] = max(pages_info["total"], pages_info["current"])
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#94A3B8"))
        canvas.drawString(2 * cm, 1.5 * cm, f"Page {pages_info['current']}")
        canvas.drawRightString(
            A4[0] - 2 * cm, 1.5 * cm, f"Generated: {now}"
        )
        canvas.restoreState()

    doc.build(
        _build_pdf_story(df, filters, full_df, insights, now, styles, title_style, heading1_style, heading2_style, body_style, caption_style),
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )

    return buffer.getvalue()


def generate_excel_report(
    df: pd.DataFrame,
    filters: dict[str, Any],
    full_df: pd.DataFrame,
) -> bytes:
    """Return a formatted Excel workbook as bytes."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "OpenPyXL is required for Excel reports. "
            "Install it with: pip install openpyxl"
        )

    buffer = io.BytesIO()
    wb = Workbook()

    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # --- Sheet 1: Raw Data ---
    ws_data = wb.active
    ws_data.title = "Filtered Dataset"

    if df.empty:
        ws_data.append(["No data available for the selected filters."])
    else:
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws_data.cell(row=1, column=c_idx, value=col.replace("_", " "))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border

        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                cell.border = cell_border
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

        for col in ws_data.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:  # noqa: BLE001
                    pass
            ws_data.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # --- Sheet 2: Summary Statistics ---
    ws_stats = wb.create_sheet(title="Summary Statistics")
    if not df.empty:
        desc = df.describe(include="all").round(2).reset_index()
        desc.rename(columns={"index": "Statistic"}, inplace=True)

        for c_idx, col in enumerate(desc.columns, 1):
            cell = ws_stats.cell(row=1, column=c_idx, value=col.replace("_", " "))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border

        for r_idx, row in enumerate(desc.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws_stats.cell(row=r_idx, column=c_idx, value=value)
                cell.border = cell_border
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
    else:
        ws_stats.append(["No data available."])

    # --- Sheet 3: KPIs ---
    ws_kpi = wb.create_sheet(title="KPI Values")
    kpis = compute_kpis(df)
    kpi_rows = [
        ["Metric", "Value"],
        ["Total Students", kpis["total_students"]],
        ["Average Exam Score", kpis["avg_exam_score"]],
        ["Average Hours Studied", f"{kpis['avg_hours_studied']} hrs/wk"],
        ["Average Attendance", f"{kpis['avg_attendance']}%"],
        ["Highest Exam Score", float(df["Exam_Score"].max()) if not df.empty else 0],
        ["Lowest Exam Score", float(df["Exam_Score"].min()) if not df.empty else 0],
    ]
    for r_idx, row in enumerate(kpi_rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_kpi.cell(row=r_idx, column=c_idx, value=value)
            cell.border = cell_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            elif r_idx % 2 == 0:
                cell.fill = alt_fill
    ws_kpi.column_dimensions["A"].width = 24
    ws_kpi.column_dimensions["B"].width = 18

    # --- Sheet 4: Correlation Matrix ---
    ws_corr = wb.create_sheet(title="Correlation Matrix")
    numeric_cols = [
        c for c in ["Hours_Studied", "Attendance", "Sleep_Hours",
                    "Previous_Scores", "Tutoring_Sessions", "Physical_Activity", "Exam_Score"]
        if c in df.columns
    ]
    if len(numeric_cols) > 1 and not df.empty:
        corr = df[numeric_cols].corr().round(2).reset_index()
        corr.rename(columns={"index": "Factor"}, inplace=True)

        for c_idx, col in enumerate(corr.columns, 1):
            cell = ws_corr.cell(row=1, column=c_idx, value=col.replace("_", " "))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border

        for r_idx, row in enumerate(corr.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws_corr.cell(row=r_idx, column=c_idx, value=value)
                cell.border = cell_border
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
    else:
        ws_corr.append(["Not enough numeric data for correlation."])

    wb.save(buffer)
    return buffer.getvalue()


def generate_csv_report(df: pd.DataFrame) -> bytes:
    """Return the filtered dataset as CSV bytes."""
    if df.empty:
        return b"No data available for the selected filters.\n"
    return df.to_csv(index=False).encode("utf-8")


def get_filter_values_from_store(store_data: dict | None) -> dict[str, Any]:
    """Normalize raw dcc.Store data into filter kwargs."""
    if not store_data:
        return {}
    return {
        "gender": store_data.get("gender", []),
        "school_type": store_data.get("school_type", []),
        "family_income": store_data.get("family_income", []),
        "parental_involvement": store_data.get("parental_involvement", []),
        "hours_studied_range": tuple(store_data["hours_studied_range"]) if store_data.get("hours_studied_range") else None,
        "attendance_range": tuple(store_data["attendance_range"]) if store_data.get("attendance_range") else None,
    }


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _build_pdf_story(df, filters, full_df, insights, now, styles, title_style, h1, h2, body, caption):
    story = []

    # ---- Cover Page ----
    story.append(Spacer(1, 4 * cm))
    story.append(Paragraph("Student Performance Dashboard", title_style))
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph("Professional Analytics Report", h2))
    story.append(Spacer(1, 2 * cm))

    cover_data = [
        ["Project:", "Student Performance Analytics Dashboard"],
        ["University:", "University Placeholder"],
        ["Author:", "Dashboard Author"],
        ["Report Date:", now],
        ["Dataset:", f"{len(full_df):,} records, {full_df.shape[1]} columns"],
    ]
    cover_table = Table(cover_data, colWidths=[4 * cm, 10 * cm])
    cover_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#334155")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(cover_table)
    story.append(PageBreak())

    # ---- Executive Summary ----
    story.append(Paragraph("Executive Summary", h1))
    kpis = compute_kpis(df)
    summary_text = (
        f"This report presents a comprehensive analysis of student performance data. "
        f"The current view includes <b>{kpis['total_students']:,} students</b> with an "
        f"average exam score of <b>{kpis['avg_exam_score']}</b>. "
        f"The report covers dataset characteristics, key performance indicators, "
        f"statistical summaries, visual trends, and actionable recommendations."
    )
    story.append(Paragraph(summary_text, body))
    story.append(Spacer(1, 0.3 * cm))

    # ---- Dataset Summary ----
    story.append(Paragraph("Dataset Summary", h1))
    missing_counts = full_df.isnull().sum()
    missing_total = int(missing_counts.sum())
    missing_cols = missing_counts[missing_counts > 0]
    missing_text = f"{missing_total} total missing values across {len(missing_cols)} columns" if missing_total > 0 else "No missing values detected"

    filter_desc = _describe_filters(filters)

    dataset_rows = [
        ["Metric", "Value"],
        ["Total Records (Filtered)", f"{len(df):,}"],
        ["Total Columns", str(df.shape[1])],
        ["Missing Values", missing_text],
        ["Active Filters", filter_desc],
    ]
    dataset_table = Table(dataset_rows, colWidths=[6 * cm, 8 * cm])
    dataset_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#334155")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(dataset_table)
    story.append(Spacer(1, 0.3 * cm))

    # ---- KPI Summary ----
    story.append(Paragraph("KPI Summary", h1))
    kpi_data = [
        ["Metric", "Value"],
        ["Total Students", f"{kpis['total_students']:,}"],
        ["Average Exam Score", f"{kpis['avg_exam_score']}"],
        ["Highest Exam Score", f"{float(df['Exam_Score'].max()):.1f}" if not df.empty else "N/A"],
        ["Lowest Exam Score", f"{float(df['Exam_Score'].min()):.1f}" if not df.empty else "N/A"],
    ]
    kpi_table = Table(kpi_data, colWidths=[6 * cm, 8 * cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#334155")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(kpi_table)
    story.append(PageBreak())

    # ---- Statistical Summary ----
    story.append(Paragraph("Statistical Summary", h1))
    if not df.empty:
        numeric_df = df.select_dtypes(include="number")
        if not numeric_df.empty:
            desc = numeric_df.describe().round(2).reset_index()
            desc.rename(columns={"index": "Statistic"}, inplace=True)

            stat_rows = [desc.columns.tolist()]
            for row in desc.itertuples(index=False):
                stat_rows.append([str(v) for v in row])

            stat_table = Table(stat_rows, repeatRows=1)
            stat_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]))
            story.append(stat_table)
        else:
            story.append(Paragraph("No numeric columns available for statistical summary.", body))
    else:
        story.append(Paragraph("No data available.", body))
    story.append(Spacer(1, 0.3 * cm))

    # ---- Charts ----
    story.append(Paragraph("Charts", h1))
    charts = _generate_charts(df)
    chart_names = {
        "distribution": "Exam Score Distribution",
        "scatter": "Hours Studied vs. Exam Score",
        "involvement": "Average Score by Parental Involvement",
        "income": "Average Score by Family Income",
    }
    for key, title in chart_names.items():
        if key not in charts:
            continue
        fig = charts[key]
        try:
            img_bytes = to_image(fig, format="png", engine="kaleido", width=800, height=500)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp.write(img_bytes)
                tmp_path = tmp.name
            img = RLImage(tmp_path, width=14 * cm, height=8.75 * cm)
            story.append(img)
            story.append(Paragraph(f"Figure: {title}", caption))
            story.append(Spacer(1, 0.4 * cm))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Failed to embed chart %s: %s", key, exc)
            story.append(Paragraph(f"[Chart '{title}' could not be embedded: {exc}]", body))
    story.append(PageBreak())

    # ---- Automatic Insights ----
    story.append(Paragraph("Automatic Insights", h1))
    if insights:
        for item in insights:
            story.append(Paragraph(f"• {item}", body))
    else:
        story.append(Paragraph("Not enough data to generate insights.", body))
    story.append(Spacer(1, 0.3 * cm))

    # ---- Conclusions ----
    story.append(Paragraph("Conclusions", h1))
    conclusions = _generate_conclusions(df, kpis)
    for item in conclusions:
        story.append(Paragraph(f"• {item}", body))
    story.append(Spacer(1, 0.3 * cm))

    # ---- Recommendations ----
    story.append(Paragraph("Recommendations", h1))
    recommendations = _generate_recommendations(df, kpis)
    for item in recommendations:
        story.append(Paragraph(f"• {item}", body))
    story.append(PageBreak())

    # ---- Footer note ----
    story.append(Paragraph("Report Footer", h1))
    story.append(Paragraph(
        f"This report was generated automatically on <b>{now}</b> using the "
        f"Student Performance Dashboard. Data source: StudentPerformanceFactors.csv. "
        f"Report contains {len(df):,} records after applying active filters. "
        f"All statistics are computed from the filtered dataset shown above.",
        body
    ))

    return story


def _describe_filters(filters: dict[str, Any]) -> str:
    parts = []
    if filters.get("gender"):
        parts.append(f"Gender: {', '.join(filters['gender'])}")
    if filters.get("school_type"):
        parts.append(f"School Type: {', '.join(filters['school_type'])}")
    if filters.get("family_income"):
        parts.append(f"Family Income: {', '.join(filters['family_income'])}")
    if filters.get("parental_involvement"):
        parts.append(f"Parental Involvement: {', '.join(filters['parental_involvement'])}")
    if filters.get("hours_studied_range"):
        lo, hi = filters["hours_studied_range"]
        parts.append(f"Hours Studied: {lo}-{hi}")
    if filters.get("attendance_range"):
        lo, hi = filters["attendance_range"]
        parts.append(f"Attendance: {lo}%-{hi}%")
    return "; ".join(parts) if parts else "None (full dataset)"


def _generate_charts(df: pd.DataFrame) -> dict[str, go.Figure]:
    """Generate the standard dashboard charts from a DataFrame."""
    if df.empty:
        return {}

    charts = {}

    fig_dist = px.histogram(df, x="Exam_Score", nbins=25, labels={"Exam_Score": "Exam Score"})
    fig_dist.update_traces(marker_line_width=0)
    apply_theme(fig_dist, title="Exam Score Distribution")
    charts["distribution"] = fig_dist

    fig_scatter = px.scatter(
        df,
        x="Hours_Studied",
        y="Exam_Score",
        color="Parental_Involvement",
        opacity=0.6,
        labels={"Hours_Studied": "Hours Studied / week", "Exam_Score": "Exam Score"},
        hover_data=["Attendance", "Sleep_Hours"],
    )
    apply_theme(fig_scatter, title="Hours Studied vs. Exam Score")
    charts["scatter"] = fig_scatter

    by_involvement = (
        df.groupby("Parental_Involvement", observed=True)["Exam_Score"]
        .mean()
        .reindex(["Low", "Medium", "High"])
        .dropna()
        .reset_index()
    )
    fig_involvement = px.bar(
        by_involvement,
        x="Parental_Involvement",
        y="Exam_Score",
        text_auto=".1f",
        labels={"Parental_Involvement": "Parental Involvement", "Exam_Score": "Avg. Exam Score"},
    )
    apply_theme(fig_involvement, title="Average Score by Parental Involvement")
    charts["involvement"] = fig_involvement

    by_income = (
        df.groupby("Family_Income", observed=True)["Exam_Score"]
        .mean()
        .reindex(["Low", "Medium", "High"])
        .dropna()
        .reset_index()
    )
    fig_income = px.bar(
        by_income,
        x="Family_Income",
        y="Exam_Score",
        text_auto=".1f",
        color_discrete_sequence=["#22C55E"],
        labels={"Family_Income": "Family Income", "Exam_Score": "Avg. Exam Score"},
    )
    apply_theme(fig_income, title="Average Score by Family Income")
    charts["income"] = fig_income

    return charts


def build_report_insights(df: pd.DataFrame) -> list[str]:
    """Return plain-text insight strings for reports."""
    if df.empty:
        return ["Not enough data to generate insights."]

    items = []

    if "Hours_Studied" in df.columns and "Exam_Score" in df.columns:
        corr = df[["Hours_Studied", "Exam_Score"]].corr().iloc[0, 1]
        direction = "positively" if corr > 0 else "negatively"
        items.append(
            f"Hours studied is {direction} correlated with exam score (r = {corr:.2f})."
        )

    if "Parental_Involvement" in df.columns:
        by_inv = df.groupby("Parental_Involvement", observed=True)["Exam_Score"].mean()
        if len(by_inv) > 1:
            best = by_inv.idxmax()
            worst = by_inv.idxmin()
            gap = by_inv.max() - by_inv.min()
            items.append(
                f"Students with '{best}' parental involvement score "
                f"{gap:.1f} points higher on average than those with '{worst}'."
            )

    if "Attendance" in df.columns:
        top_attendance = df["Attendance"].quantile(0.75)
        high_att_avg = df.loc[df["Attendance"] >= top_attendance, "Exam_Score"].mean()
        low_att_avg = df.loc[df["Attendance"] < top_attendance, "Exam_Score"].mean()
        if pd.notna(high_att_avg) and pd.notna(low_att_avg):
            items.append(
                f"Students in the top attendance quartile average "
                f"{high_att_avg - low_att_avg:.1f} points higher than the rest."
            )

    if (df["Exam_Score"] > 95).any():
        n_top = int((df["Exam_Score"] > 95).sum())
        items.append(
            f"{n_top} student(s) scored above 95, suggesting a small "
            f"high-performing outlier group worth investigating."
        )

    return items if items else ["Not enough data to generate insights."]


def _generate_conclusions(df: pd.DataFrame, kpis: dict) -> list[str]:
    """Generate conclusion statements from the filtered data."""
    if df.empty:
        return ["No data available to draw conclusions."]

    conclusions = []

    if kpis["avg_exam_score"] >= 80:
        conclusions.append(
            "The filtered cohort demonstrates strong overall academic performance, "
            "with an average exam score above 80."
        )
    elif kpis["avg_exam_score"] >= 65:
        conclusions.append(
            "The filtered cohort shows moderate academic performance, with an "
            "average exam score in the 65-80 range."
        )
    else:
        conclusions.append(
            "The filtered cohort shows areas for improvement, with an average "
            "exam score below 65."
        )

    if "Hours_Studied" in df.columns and "Exam_Score" in df.columns:
        corr = df[["Hours_Studied", "Exam_Score"]].corr().iloc[0, 1]
        if abs(corr) > 0.3:
            direction = "positive" if corr > 0 else "negative"
            conclusions.append(
                f"A statistically meaningful {direction} correlation (r={corr:.2f}) exists "
                f"between hours studied and exam score."
            )

    if "Attendance" in df.columns and "Exam_Score" in df.columns:
        top_att = df["Attendance"].quantile(0.75)
        high_avg = df.loc[df["Attendance"] >= top_att, "Exam_Score"].mean()
        low_avg = df.loc[df["Attendance"] < top_att, "Exam_Score"].mean()
        if pd.notna(high_avg) and pd.notna(low_avg) and (high_avg - low_avg) > 3:
            conclusions.append(
                "Students with higher attendance consistently outperform their peers, "
                "suggesting attendance is a key performance driver."
            )

    if "Parental_Involvement" in df.columns:
        by_inv = df.groupby("Parental_Involvement", observed=True)["Exam_Score"].mean()
        if len(by_inv) > 1:
            conclusions.append(
                "Parental involvement level is associated with meaningful differences "
                "in average exam scores across the cohort."
            )

    if len(df) < 100:
        conclusions.append(
            "The filtered dataset is relatively small; conclusions should be "
            "interpreted with caution."
        )

    return conclusions if conclusions else ["Analysis complete. Review the detailed statistics above."]


def _generate_recommendations(df: pd.DataFrame, kpis: dict) -> list[str]:
    """Generate actionable recommendations from the filtered data."""
    if df.empty:
        return ["No data available to generate recommendations."]

    recs = []

    if kpis["avg_hours_studied"] < 20:
        recs.append(
            "Encourage increased study time. The cohort averages less than 20 hours/week, "
            "which may limit performance potential."
        )

    if kpis["avg_attendance"] < 80:
        recs.append(
            "Improve attendance rates. The current average attendance is below 80%, "
            "which likely impacts learning outcomes."
        )

    if "Parental_Involvement" in df.columns:
        by_inv = df.groupby("Parental_Involvement", observed=True)["Exam_Score"].mean()
        if "Low" in by_inv.index and by_inv.loc["Low"] < by_inv.mean():
            recs.append(
                "Launch targeted parental engagement programs for families with low "
                "involvement, as this group shows lower average scores."
            )

    if "Sleep_Hours" in df.columns:
        avg_sleep = df["Sleep_Hours"].mean()
        if avg_sleep < 7:
            recs.append(
                f"Promote healthy sleep habits. The cohort averages {avg_sleep:.1f} hours "
                f"of sleep, below the recommended 7-9 hours for optimal cognitive function."
            )

    recs.append(
        "Monitor progress with regular assessments and adjust interventions based on "
        "measured outcomes."
    )

    return recs
